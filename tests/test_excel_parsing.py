"""Tests for the Excel parser's column detection and the
output path layout.

The previous implementation had a critical bug: the keyword
lists for "no" and "desc" overlapped (both contained the
Arabic word "بند"), so the row-number column was being claimed
by the description-column matcher. The output Excel then put
the row number into the description slot and the real Arabic
description text disappeared (was never read).

These tests pin the contract down: for a real Arabic BOQ
header (`بند | بيان الأعمال | الوحدة | الكمية | الفئة | الإجمالي`),
"بند" maps to "no" and "بيان الأعمال" maps to "desc", and the
actual data rows have the Arabic description text in the
"Item Description" field (not the row number).
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
import pytest

# conftest.py handles the project-root sys.path insertion and the
# Qt offscreen platform setup. See tests/conftest.py.
from core import excel  # noqa: E402


# ---------------------------------------------------------------------------
# detect_columns
# ---------------------------------------------------------------------------


def test_detect_arabic_boq_header():
    """A real Arabic BOQ header: بند|بيان الأعمال|الوحدة|الكمية|الفئة|الإجمالي

    Each column should map to its correct label. The critical
    bug-fix assertion is that "بند" maps to "no", not to "desc"
    (the previous code's `desc_kw` list contained "بند" which
    would steal this column from the row-number matcher).
    """
    headers = ["بند", "بيان الأعمال", "الوحدة", "الكمية", "الفئة", "الإجمالي"]
    cols = excel.detect_columns(headers)
    assert cols.get("no") == 0
    assert cols.get("desc") == 1
    assert cols.get("unit") == 2
    assert cols.get("qty") == 3
    assert cols.get("rate") == 4
    assert cols.get("total") == 5


def test_detect_english_boq_header():
    headers = ["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"]
    cols = excel.detect_columns(headers)
    assert cols.get("no") == 0
    assert cols.get("desc") == 1
    assert cols.get("unit") == 2
    assert cols.get("qty") == 3
    assert cols.get("rate") == 4
    assert cols.get("total") == 5


def test_detect_partial_header_just_nr_and_desc():
    """A BOQ with only Nr + Description is still a valid sheet."""
    headers = ["Nr.", "Description"]
    cols = excel.detect_columns(headers)
    assert cols.get("no") == 0
    assert cols.get("desc") == 1


def test_detect_arabic_partial_header():
    """Arabic equivalent of just Nr + Description."""
    headers = ["بند", "البيان"]
    cols = excel.detect_columns(headers)
    assert cols.get("no") == 0
    assert cols.get("desc") == 1


def test_detect_no_match_returns_empty():
    """A header that doesn't match any keywords returns an empty
    mapping (caller is expected to skip such sheets)."""
    headers = ["Foo", "Bar", "Baz"]
    assert excel.detect_columns(headers) == {}


# ---------------------------------------------------------------------------
# parse_excel on a real Arabic BOQ
# ---------------------------------------------------------------------------


# Real BOQ used by the end-to-end regression test. Opt-in via the
# TAWREED_TEST_BOQ env var so we don't commit a path that doesn't
# exist on every developer's machine. Set in your shell before
# running pytest if you have a real BOQ to test against.
REAL_BOQ = Path(os.environ.get("TAWREED_TEST_BOQ", "")) if os.environ.get("TAWREED_TEST_BOQ") else None


@pytest.mark.skipif(not (REAL_BOQ and REAL_BOQ.is_file()),
                    reason="Set TAWREED_TEST_BOQ to a real BOQ path to run this test")
def test_real_arabic_boq_columns_aligned():
    """End-to-end: parse the real file, verify each row's data ends
    up in the right cell. This is the regression test for the
    'Item Description' containing the row number bug."""
    md, data, headers = excel.parse_excel(str(REAL_BOQ))
    assert data, "No data was extracted from the real BOQ"
    # At least one sheet should have been processed.
    assert headers, "No sheet headers were identified"

    # Find at least one row with a non-trivial Arabic description.
    arabic_desc_found = False
    for g_id, row in data.items():
        desc = row.get("Item Description", "")
        if any("\u0600" <= c <= "\u06FF" for c in desc) and len(desc) > 5:
            # The description must NOT be just a row number.
            assert not desc.strip().isdigit(), (
                f"Row {g_id}: Arabic description slot contains a number "
                f"({desc!r}) — column misalignment regression"
            )
            arabic_desc_found = True
            break
    assert arabic_desc_found, (
        "No Arabic description text was extracted — every row's "
        "description slot is empty or numeric"
    )


# ---------------------------------------------------------------------------
# Output path layout — everything under ~/.tawreed
# ---------------------------------------------------------------------------


def test_output_path_is_under_home_tawreed(monkeypatch):
    """Per project policy, state lives at ``~/.tawreed/`` regardless
    of dev vs frozen mode. The path must NOT depend on the EXE
    location, the project root, or %LOCALAPPDATA%."""
    # Point expanduser at a tmp dir so we don't pollute the real home.
    fake_home = Path(os.environ.get("TEMP", "/tmp")) / "fake_home_tawreed_test"
    monkeypatch.setattr(os.path.expanduser, "__defaults__", None, raising=False)
    monkeypatch.setattr(os.path, "expanduser", lambda p: str(fake_home) if p == "~" else p)

    # Reload core.db so it picks up the new expanduser.
    import importlib
    import core.db as db
    importlib.reload(db)

    assert db.TAWREED_DIR == str(fake_home / ".tawreed")
    assert db.OUTPUTS_DIR == str(fake_home / ".tawreed" / "outputs")
    assert db.DB_PATH == str(fake_home / ".tawreed" / "db" / "tawreed.db")
    assert db.CONFIG_PATH == str(fake_home / ".tawreed" / "config.json")
    assert db.PID_FILE_PATH == str(fake_home / ".tawreed" / "single-instance.pid")
    # Defensive: TAWREED_DIR must be `<fake_home>/.tawreed` — not
    # embedded in the EXE path, not under %LOCALAPPDATA%, not under
    # the project root. The fake_home itself is in TEMP for
    # isolation, but the layout still proves the resolution works.
    assert db.TAWREED_DIR.endswith(os.path.join(".tawreed"))
    # The path is anchored to the user home, not the EXE dir.
    assert str(fake_home) in db.TAWREED_DIR
    assert "dist" not in db.TAWREED_DIR.split(os.sep)
    # No "Tawreed" prefix from the EXE folder either.
    assert "Tawreed.exe" not in db.TAWREED_DIR


def test_output_path_is_identical_for_frozen_and_dev(monkeypatch):
    """Frozen mode and dev mode must produce the same path. The
    user wants one canonical location, not two."""
    fake_home = Path(os.environ.get("TEMP", "/tmp")) / "fake_home_tawreed_frozen"
    monkeypatch.setattr(os.path, "expanduser", lambda p: str(fake_home) if p == "~" else p)

    # Simulate frozen: sys.frozen=True, sys.executable somewhere weird.
    fake_exe = fake_home / "Desktop" / "QS Mind" / "tawreed-api" / "dist" / "Tawreed" / "Tawreed.exe"
    fake_exe.parent.mkdir(parents=True, exist_ok=True)
    fake_exe.write_text("")
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(fake_exe), raising=False)

    import importlib
    import core.db as db_frozen
    importlib.reload(db_frozen)
    frozen_path = db_frozen.TAWREED_DIR

    # Now switch to dev mode (frozen=False) and reload.
    monkeypatch.setattr(sys, "frozen", False, raising=False)
    importlib.reload(db_frozen)
    dev_path = db_frozen.TAWREED_DIR

    # Both must point at the SAME place.
    assert os.path.normcase(frozen_path) == os.path.normcase(dev_path)
    assert frozen_path == str(fake_home / ".tawreed")


# ---------------------------------------------------------------------------
# Roundtrip: write Excel and read it back
# ---------------------------------------------------------------------------


def test_write_then_read_preserves_arabic(tmp_path):
    """A roundtrip through write_excel preserves the Arabic text
    in the description column (no cp437/cp1252 mojibake regression)."""
    from core.excel import write_excel, parse_excel
    out = tmp_path / "out.xlsx"
    row_mapping = {
        "R1": {
            "Nr.": "1",
            "Item Description": "نظام الإنذار المعنون ضد الحريق",
            "Unit": "نظام",
            "Qty": 1.0,
            "Rate": 1000.0,
            "Amount": 0,
            "sheet_name": "Test",
            "original_values": {
                "Nr.": "1", "nr": "1",
                "Item Description": "نظام الإنذار المعنون ضد الحريق",
                "description": "نظام الإنذار المعنون ضد الحريق",
                "Unit": "نظام", "unit": "نظام",
                "Qty": 1.0, "qty": 1.0,
                "Rate": 1000.0, "rate": 1000.0,
                "Amount": 0, "amount": 0,
            },
        },
    }
    item_categories = {"R1": "Fire Alarm"}
    write_excel(str(out), row_mapping, item_categories,
                project_name="اختبار", date="2026-06-13")

    # Read back and verify the description cell holds Arabic.
    wb = openpyxl.load_workbook(str(out), data_only=False)
    pkg = next(s for s in wb.sheetnames if s.startswith("Pkg"))
    ws = wb[pkg]
    desc_cell = ws.cell(row=2, column=2).value
    assert "نظام الإنذار" in (desc_cell or ""), (
        f"Arabic description did not survive the roundtrip; got {desc_cell!r}"
    )


def test_write_excel_caps_description_column_width(tmp_path):
    """Regression: the old code auto-fit every column to the
    longest cell's character count, which meant a single 800-char
    Arabic description widened the description column to ~810.
    The new code caps the column at the user-requested width
    (60) so the spreadsheet actually looks usable."""
    from core.excel import write_excel, COL_WIDTH_DESC
    out = tmp_path / "caps.xlsx"
    long_arabic = "وصف طويل " * 200  # ~2000 chars
    row_mapping = {
        "R1": {
            "Nr.": "1",
            "Item Description": long_arabic,
            "Unit": "نظام",
            "Qty": 1.0,
            "Rate": 1000.0,
            "Amount": 0,
            "original_values": {
                "Nr.": "1", "nr": "1",
                "Item Description": long_arabic,
                "description": long_arabic,
                "Unit": "نظام", "unit": "نظام",
                "Qty": 1.0, "qty": 1.0,
                "Rate": 1000.0, "rate": 1000.0,
                "Amount": 0, "amount": 0,
            },
        },
    }
    item_categories = {"R1": "Test"}
    write_excel(str(out), row_mapping, item_categories,
                project_name="اختبار", date="2026-06-13")

    wb = openpyxl.load_workbook(str(out), data_only=False)
    pkg = next(s for s in wb.sheetnames if s.startswith("Pkg"))
    ws = wb[pkg]
    desc_width = ws.column_dimensions["B"].width
    assert desc_width <= COL_WIDTH_DESC, (
        f"Description column width {desc_width} exceeds the {COL_WIDTH_DESC} "
        "cap — the long-cell blowup regression is back."
    )


def test_write_excel_freezes_header_row(tmp_path):
    """Frozen panes must be set to A2 so the header row stays
    visible while scrolling."""
    from core.excel import write_excel
    out = tmp_path / "frozen.xlsx"
    row_mapping = {
        "R1": {
            "Nr.": "1",
            "Item Description": "x",
            "Unit": "u",
            "Qty": 1.0,
            "Rate": 1.0,
            "Amount": 0,
            "original_values": {
                "Nr.": "1", "nr": "1",
                "Item Description": "x", "description": "x",
                "Unit": "u", "unit": "u",
                "Qty": 1.0, "qty": 1.0,
                "Rate": 1.0, "rate": 1.0,
                "Amount": 0, "amount": 0,
            },
        },
    }
    item_categories = {"R1": "Cat"}
    write_excel(str(out), row_mapping, item_categories,
                project_name="p", date="2026-01-01")
    wb = openpyxl.load_workbook(str(out), data_only=False)
    pkg = next(s for s in wb.sheetnames if s.startswith("Pkg"))
    ws = wb[pkg]
    assert ws.freeze_panes == "A2"


def test_write_excel_amount_is_formula(tmp_path):
    """The Amount cell must be a formula (=D*E), not a hardcoded
    value. Per Anthropic xlsx skill: formulas, not hardcodes."""
    from core.excel import write_excel
    out = tmp_path / "formula.xlsx"
    row_mapping = {
        "R1": {
            "Nr.": "1",
            "Item Description": "x",
            "Unit": "u",
            "Qty": 2.0,
            "Rate": 3.5,
            "Amount": 0,
            "original_values": {
                "Nr.": "1", "nr": "1",
                "Item Description": "x", "description": "x",
                "Unit": "u", "unit": "u",
                "Qty": 2.0, "qty": 2.0,
                "Rate": 3.5, "rate": 3.5,
                "Amount": 0, "amount": 0,
            },
        },
    }
    item_categories = {"R1": "Cat"}
    write_excel(str(out), row_mapping, item_categories,
                project_name="p", date="2026-01-01")
    wb = openpyxl.load_workbook(str(out), data_only=False)
    pkg = next(s for s in wb.sheetnames if s.startswith("Pkg"))
    ws = wb[pkg]
    amount_cell = ws.cell(row=2, column=6).value
    assert isinstance(amount_cell, str) and amount_cell.startswith("="), (
        f"Amount cell should be a formula like =IFERROR(D2*E2, 0); got {amount_cell!r}"
    )
