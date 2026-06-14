"""End-to-end tests for the BOQ processing pipeline.

These tests pin down the contract that connects:
  - ``core.excel.parse_excel`` (input parser)
  - ``core.ai.analyze_boq_stream`` (LLM categoriser)
  - ``core.excel.write_excel`` (output writer)
  - ``core.db.add_history`` (history row)

The LLM is fully mocked — no network calls, no real API key.
The mock streams the JSON the real ``analyze_boq_stream`` would
have produced, so the test exercises the actual production code
paths (JSON parsing, sentinel detection, the flat-dict-to-items-
lift, the row-mapping back to the input rows, and the
``write_excel`` happy path).

The point isn't to retest the LLM SDK; it's to catch regressions
in the *glue* — the order of operations, the arguments passed
between functions, and the error envelope that's emitted when
the LLM returns garbage.
"""
from __future__ import annotations

from pathlib import Path
from unittest import mock

import openpyxl
import pytest


@pytest.fixture(autouse=True)
def _init_db_for_e2e(isolated_tawreed_dir):
    """Create the schema before each test runs. The
    ``isolated_tawreed_dir`` fixture points ``core.db`` at a
    tmp dir + neutralises legacy detection; this autouse fixture
    just calls ``init_db()`` so ``add_history`` and
    ``get_history`` have a real ``history`` table to talk to."""
    from core import db
    db.init_db()


# A tiny but realistic BOQ — 3 items, English headers. Built in
# memory so we don't depend on a file existing on disk.
def _build_input_xlsx(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append(["1", "Excavation", "m3", 100, 50, 5000])
    ws.append(["2", "Concrete C30", "m3", 50, 800, 40000])
    ws.append(["3", "Rebar 16mm", "ton", 2, 4500, 9000])
    wb.save(path)
    return path


def _fake_openai_stream(chunks: list):
    """Build a fake openai streaming response that yields the
    given chunks as ``choices[0].delta.content``."""
    class _Chunk:
        def __init__(self, content):
            self.choices = [
                type(
                    "C", (),
                    {"delta": type("D", (), {"content": content, "reasoning_content": None})()},
                )()
            ]

    class _Resp:
        def __iter__(self):
            for c in chunks:
                yield _Chunk(c)

    return _Resp()


def _patch_openai_with_stream(monkeypatch, json_chunks: list) -> None:
    """Replace ``openai.OpenAI`` with a stub that streams the
    given JSON chunks. Mirrors the helper in test_streaming.py —
    kept independent to avoid a brittle cross-test import."""
    from core import ai

    fake_resp = _fake_openai_stream(json_chunks)
    monkeypatch.setattr(ai.openai.OpenAI, "__init__", lambda self, **kw: None)
    chat_mock = mock.Mock(**{"completions.create.return_value": fake_resp})
    monkeypatch.setattr(ai.openai.OpenAI, "chat", chat_mock)


def _drain_stream(gen):
    """Drain ``analyze_boq_stream`` to completion. Returns the
    final parsed dict (the payload of the last yield, which is
    the ``("__DONE__", parsed_data)`` tuple)."""
    final = None
    for event in gen:
        # The generator yields (text, is_thought) for tokens, and
        # the terminal yield is ("__DONE__", parsed_data).
        if isinstance(event, tuple) and len(event) == 2 and event[0] == "__DONE__":
            final = event[1]
            break
    return final


def test_e2e_processing_pipeline_with_valid_boq(monkeypatch, tmp_path, isolated_tawreed_dir):
    """End-to-end: parse input -> categorise via mocked LLM ->
    write output -> record history. The output file exists, has
    the expected sheets, and the history row was added."""
    from core import ai, db, excel

    # 1. Build a real input file.
    input_path = _build_input_xlsx(tmp_path / "in.xlsx")
    output_path = tmp_path / "out.xlsx"

    # 2. Parse it.
    md, data, _headers = excel.parse_excel(str(input_path))
    assert data, "parser returned no data"
    assert len(data) >= 3, f"expected at least 3 rows, got {len(data)}"

    # 3. Mock the LLM and run the categoriser.
    flat_json = (
        '{"R1": "Earthwork", "R2": "Concrete", "R3": "Steel", '
        '"project_name": "Test", "date": "2026-06-14"}'
    )
    _patch_openai_with_stream(monkeypatch, [flat_json])

    final = _drain_stream(
        ai.analyze_boq_stream(
            md, "OpenAI", "sk-test-fake", "gpt-4o-mini", ""
        )
    )
    assert final is not None, "analyzer did not yield __DONE__ terminal"
    assert final.get("project_name") == "Test"

    # 4. Wire the flat dict back to items-key shape and write Excel.
    items_payload = final.get("items", final)
    row_mapping = {}
    item_categories = {}
    for idx, (row_id, row) in enumerate(sorted(data.items()), start=1):
        row_mapping[row_id] = row
        key = f"R{idx}"
        item_categories[row_id] = items_payload.get(key, "General")

    excel.write_excel(
        str(output_path), row_mapping, item_categories, "Test", "2026-06-14"
    )
    assert output_path.exists(), "output Excel was not written"
    assert output_path.stat().st_size > 0, "output Excel is empty"

    # 5. Verify the output structure.
    out_wb = openpyxl.load_workbook(str(output_path), data_only=True)
    assert "Cover" in out_wb.sheetnames, f"missing Cover; got {out_wb.sheetnames}"
    assert "Master" in out_wb.sheetnames, f"missing Master; got {out_wb.sheetnames}"
    pkg_sheets = [n for n in out_wb.sheetnames if n.startswith("Pkg - ")]
    assert pkg_sheets, f"no package sheets; got {out_wb.sheetnames}"

    # 6. Record history (should not raise).
    db.add_history("Test", len(set(item_categories.values())), str(output_path))
    rows = db.get_history()
    assert len(rows) == 1, f"expected 1 history row, got {len(rows)}"
    assert rows[0]["project_name"] == "Test"


def test_e2e_handles_garbage_llm_response_gracefully(monkeypatch, tmp_path, isolated_tawreed_dir):
    """If the LLM returns something that isn't valid JSON, the
    analyzer should still emit __DONE__ with an error envelope —
    never raise."""
    from core import ai

    _patch_openai_with_stream(monkeypatch, ["this is not json {{{"])

    final = _drain_stream(
        ai.analyze_boq_stream(
            "fake markdown", "OpenAI", "sk-test-fake", "gpt-4o-mini", ""
        )
    )
    assert final is not None, "analyzer did not yield __DONE__ on garbage LLM response"
    # The error envelope should be flagged — at minimum via an
    # "error" key OR an empty / fallback items dict.
    assert "error" in final or not final.get("items"), (
        f"expected an error envelope or empty items; got {final}"
    )


def test_e2e_roundtrip_preserves_amount_formulas(monkeypatch, tmp_path, isolated_tawreed_dir):
    """The Amount column is a formula (=D*E or =IFERROR(D*E,0)).
    Verify the round-trip writes a formula (not a literal value)
    in the Amount column of every row, with the right row
    references to Qty (D) and Rate (E)."""
    from core import ai, excel

    input_path = _build_input_xlsx(tmp_path / "in.xlsx")
    md, data, _headers = excel.parse_excel(str(input_path))

    flat_json = (
        '{"R1": "General", "R2": "General", "R3": "General", '
        '"project_name": "Formulas", "date": "2026-06-14"}'
    )
    _patch_openai_with_stream(monkeypatch, [flat_json])

    final = _drain_stream(
        ai.analyze_boq_stream(
            md, "OpenAI", "sk-test-fake", "gpt-4o-mini", ""
        )
    )
    assert final is not None

    items_payload = final.get("items", final)
    row_mapping = {}
    item_categories = {}
    for idx, (row_id, row) in enumerate(sorted(data.items()), start=1):
        row_mapping[row_id] = row
        item_categories[row_id] = items_payload.get(f"R{idx}", "General")

    output_path = tmp_path / "formulas.xlsx"
    excel.write_excel(str(output_path), row_mapping, item_categories, "Formulas", "2026-06-14")

    # Reopen WITHOUT data_only=True so we see the formula text
    # (openpyxl in data_only mode returns cached values, which
    # aren't computed for files it just wrote).
    out_wb = openpyxl.load_workbook(str(output_path), data_only=False)
    master = out_wb["Master"]
    headers = [cell.value for cell in master[1]]
    assert "Amount" in headers, f"Amount column not in headers: {headers}"
    amount_col = headers.index("Amount") + 1

    # Every data row's Amount cell should be a formula starting
    # with "=". The exact form varies (we tested for =IFERROR(D*E,0)
    # in test_excel_parsing.py, so any formula here is fine).
    formula_count = 0
    for row in master.iter_rows(min_row=2):
        cell = row[amount_col - 1]
        if cell.value is not None:
            value = str(cell.value)
            assert value.startswith("="), (
                f"Amount cell at {cell.coordinate} is not a formula: {value!r}"
            )
            # Must reference Qty and Rate (columns D and E in our
            # schema) — that's the whole point of the formula.
            assert "D" in value and "E" in value, (
                f"Amount formula at {cell.coordinate} does not reference D*E: {value!r}"
            )
            formula_count += 1
    assert formula_count == 3, f"expected 3 amount formulas, got {formula_count}"


def test_e2e_history_records_multiple_runs(tmp_path, isolated_tawreed_dir):
    """Two consecutive runs should produce two history rows, in
    the order they were added (or reverse, depending on the
    query). Pins down the contract that history persists across
    runs in the same ~/.tawreed/."""
    from core import db

    db.add_history("Project A", 3, str(tmp_path / "a.xlsx"))
    db.add_history("Project B", 5, str(tmp_path / "b.xlsx"))
    rows = db.get_history()
    assert len(rows) == 2
    names = {r["project_name"] for r in rows}
    assert names == {"Project A", "Project B"}
