import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample():
    # Define data for the mock BOQ
    data = {
        "Item No.": ["1.1", "1.2", "2.1", "2.2", "2.3", "3.1", "3.2"],
        "Description": [
            "Supply and install 60x60cm high-quality porcelain floor tiles, color light beige, including mortar bed, grouting, protective sheet, and cleaning. To be selected by consultant.",
            "Supply and install solid teak wood skirting 10cm height, satin lacquer finish, including all fixings and accessories.",
            "Supply and install solid core timber doors, size 900x2200mm, with walnut veneer finish, including hardwood frame, architrave, heavy-duty stainless steel hinges, mortise lockset, and door closer.",
            "Supply and install glazed aluminum double doors, size 1800x2200mm, powder coated finish, 12mm clear tempered glass, with floor spring, handles, and locks.",
            "Supply and install stainless steel push plates and pull handles for doors as per schedule.",
            "Supply and install gypsum board suspended ceiling, 12mm thickness, moisture resistant for wet areas, including metal framing, joint tape, sanding, and 3 coats of emulsion paint.",
            "Supply and install 60x60cm mineral fiber tile lay-in ceiling with exposed T-grid system, including all hangers and wall angles."
        ],
        "Unit": ["m2", "m", "No", "No", "set", "m2", "m2"],
        "Qty": [120, 85, 12, 4, 16, 140, 95],
        "Rate": [150, 45, 1200, 2500, 180, 95, 75],
        "Amount": [18000, 3825, 14400, 10000, 2880, 13300, 7125]
    }

    df = pd.DataFrame(data)

    # Save to Excel
    file_path = "sample_boq.xlsx"
    
    # We will use openpyxl to make it look like a typical client BOQ
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fit-out BOQ"
    ws.views.sheetView[0].showGridLines = True

    # Title Row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "BILL OF QUANTITIES - FIT-OUT PROJECT"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 40

    # Headers
    headers = list(df.columns)
    for col_idx, h in enumerate(headers):
        cell = ws.cell(row=3, column=col_idx + 1, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 24

    border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Write Data
    for r_idx, row in df.iterrows():
        row_num = r_idx + 4
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=row_num, column=c_idx + 1, value=val)
            cell.border = data_border
            cell.font = Font(name="Calibri", size=10)
            
            # Alignments
            if c_idx == 0 or c_idx == 2:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 1:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="right")
                
            # Number formats
            if c_idx in [3, 4, 5]:
                try:
                    cell.value = float(val)
                    if c_idx == 3:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.00"
                except ValueError:
                    pass
        ws.row_dimensions[row_num].height = 36

    # Column Widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

    # Save
    wb.save(file_path)
    print(f"Created {file_path}")

if __name__ == "__main__":
    create_sample()
