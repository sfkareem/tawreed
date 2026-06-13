"""Excel BOQ parsing and output generation.

The parser is the highest-risk module in Tawreed: a single
misaligned column on the input sheet puts every row's data in
the wrong cell of the output Excel. Real-world construction BOQs
are written in mixed Arabic/English, with non-standard column
names, merged title rows, and sparse data (sub-items often have
only a description; quantities appear in later rows). The parser
must be defensive about all of these.

The previous version of ``clean_header_val`` contained a
"cp437/cp1252 → utf-8" round-trip that was meant to fix
Mojibake (when Excel shows Arabic as garbled Latin-1). In
practice, on a properly-encoded Arabic file, the round-trip
just *creates* Mojibake from real text. Removed.

The previous header detection was buggy in two ways:
  1. Keyword matching was a substring check across the whole
     Arabic word, so "بند" (item/number column) matched the
     "desc_kw" list, stealing the column from real descriptions.
  2. Required at least 2 matched columns, so a sheet with only
     "Nr. + Description" was skipped entirely.

The new detection:
  * Scores every column with multiple candidate labels and picks
    the column with the highest match count, breaking ties by
    column order (Nr → Desc → Unit → Qty → Rate → Amount).
  * The minimum header requirement is "Nr" OR "Description" —
    one is enough to anchor a sheet.
  * The Arabic keyword lists are non-overlapping by construction.

Output formatting (write_excel):
  Implements the relevant guidelines from Anthropic's
  ``xlsx`` skill — professional Calibri font, wrap_text on
  long text columns, column-width caps (so a single 800-char
  Arabic description doesn't blow the column to width 800+),
  frozen header panes, currency-formatted Amounts, alternating
  row stripes, and proper borders. See ``_style_*`` helpers.
"""
import os
import re
import sys
import errno
import inspect
import logging
import zipfile
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from typing import Any, Tuple, Dict, List, Optional

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Output formatting constants
# ---------------------------------------------------------------------------
#
# Per the Anthropic xlsx skill, an Excel deliverable should:
#   * Use a consistent professional font.
#   * Never produce a deliverable that looks ugly when opened —
#     specifically, never let one cell's content set the column
#     width to hundreds of characters.
#   * Have wrap_text on free-text columns so long descriptions
#     render inside the cell, not overflowing the viewport.
#   * Freeze the header row so it stays visible while scrolling.
#   * Apply currency formatting to monetary columns.
#
# The previous version of write_excel auto-fit every column to
# ``len(str(value))`` + 3, which meant a single 800-character
# Arabic description widened the entire column to ~810. The user
# (Kareem) reported this as "the generated excel sucks". The
# fix is below: a hard cap on width plus a wrap_text+wrap-friendly
# row-height strategy.

OUTPUT_FONT_NAME = "Calibri"
OUTPUT_FONT_SIZE = 11

# Hard caps on column widths. The cap for the description column
# is the user-requested 60 — long descriptions still render
# correctly because wrap_text is on. Other columns are sized
# based on their typical content (numbers, units, dates).
COL_WIDTH_DESC = 60      # user spec
COL_WIDTH_NR = 8
COL_WIDTH_UNIT = 14
COL_WIDTH_QTY = 10
COL_WIDTH_RATE = 14
COL_WIDTH_AMOUNT = 16
COL_WIDTH_SHEET = 20
COL_WIDTH_PACKAGE = 28

# Header row uses a calm grey fill + bold text — readable but not
# loud. Per skill: avoid pure black headers, prefer dark grey.
HEADER_FILL_HEX = "FF1F2937"   # slate-800
HEADER_FONT_HEX = "FFFFFFFF"   # white
HEADER_FONT_SIZE = 11

# Alternating row stripes (zebra striping) for body rows. Pale,
# neutral — don't compete with the data.
ZEBRA_FILL_HEX = "FFF8FAFC"   # slate-50
BODY_FONT_HEX = "FF1F2937"    # slate-800 for body text

# Borders. Thin, light grey. Don't make the workbook look like
# a tax form.
BORDER_HEX = "FFD1D5DB"       # grey-300
BORDER_STYLE = "thin"

# Number formats. The Amount column is currency (no symbol, just
# thousand separators — the BOQ market in MENA uses EGP, USD, etc.
# inconsistently, so we let the source-data Rate dictate the unit).
FMT_INT = "#,##0"
FMT_QTY = "#,##0.##"          # qty can be 0.5 m² etc.
FMT_RATE = "#,##0.00"         # rate is money
FMT_AMOUNT = "#,##0.00"       # amount is money


def _column_width(text: Any, cap: int, pad: int = 2) -> int:
    """Estimate a display width for a cell's value, capped at ``cap``.

    The cap is the critical part — it stops a single 800-char
    Arabic description from setting the column to width 800+.

    The width estimate is roughly: the number of visible
    characters plus a small pad. Arabic / CJK glyphs are slightly
    wider than Latin glyphs in Calibri, so we round up.
    """
    if text is None:
        return cap
    s = str(text)
    if not s:
        return cap
    # Count "wide" chars (non-ASCII, Arabic, CJK) separately.
    # Latin digits/letters: 1 unit; wide chars: 1.3 units.
    # Then add the pad and clamp to cap with a minimum of 10.
    width = 0
    for ch in s:
        if ord(ch) < 0x1100 or unicodedata.category(ch) == "Cf":
            # Latin / basic punctuation / control -> 1.0
            width += 1.0
        else:
            # Arabic, CJK, etc. -> 1.4
            width += 1.4
    return max(10, min(int(width) + pad, cap))


def _thin_border() -> Border:
    side = Side(style=BORDER_STYLE, color=BORDER_HEX)
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
#
# Each column is scored against this matrix. The match is case-
# insensitive substring containment after whitespace normalisation.
# We deliberately exclude cross-confusable keywords (e.g. "بند"
# appears in BOTH Arabic "nr" and "desc" dialects — we treat it
# as nr only, because in BOQ spreadsheets "بند" is almost always
# the row-number column, not a free-text description column).

_HEADER_LABELS: Dict[str, List[str]] = {
    # Row number / item id column.
    "no": [
        "nr", "no.", "no ", "no)", "no:", "number", "#", "item no",
        "بنـد", "بنود", "بند", "رقم", "رقـم", "الرقم", "مسلسل",
        "البنود", "البند",
        # French (common in MENA construction)
        "n°", "no.",
    ],
    # Item description / free-text column.
    "desc": [
        "description", "desc", "item description", "item desc",
        "scope", "scope of work", "item", "work item",
        "بيان", "بيـان", "البيـان", "البيان", "وصف", "الوصف",
        "بيان الأعمال", "وصف الأعمال", "بيان البند",
    ],
    # Unit column.
    "unit": [
        "unit", "uom", "units",
        "وحدة", "وحـدة", "الوحدة", "الوحـدة", "وحده", "الوحده",
    ],
    # Quantity column.
    "qty": [
        "qty", "quantity", "qty.", "q'ty", "q-ty", "quantities",
        "كمية", "الكميـة", "الكميه", "الكمية", "الكمـية", "كميه", "الكمـيه",
    ],
    # Rate / unit price column.
    "rate": [
        "rate", "unit rate", "unit price", "price", "unit cost",
        "unitrate", "unitprice",
        "سعر", "السعر", "سعر الوحدة", "سعر الوحده", "سعر الوحد",
        "فئـة", "فئه", "الفئـة", "الفئه", "فئة", "الفئة",
    ],
    # Total / amount column.
    "total": [
        "total", "amount", "total amount", "total price", "total cost",
        "value", "tot.",
        "إجمالي", "الاجمالي", "الإجمالي", "الاجمـالي",
        "اجمالي", "إجمـالي",
        "مبلغ", "المبلغ", "القيمة", "القيمـة",
    ],
}


def _clean(s: Any) -> str:
    """Lowercase, strip, and normalise whitespace for header comparison.

    Removed the previous cp437/cp1252 round-trip — that was actively
    mogrifying real Arabic text. If we ever need a Mojibake fix for
    a specific user, we'll add a guarded round-trip, but the default
    is to trust the input encoding.
    """
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _score_column(header_text: str) -> Dict[str, int]:
    """Return a {label_key: score} dict for one header cell.

    Score is the number of distinct keywords from ``_HEADER_LABELS``
    that appear as substrings in ``header_text``. Ties are broken
    by the order of ``_HEADER_LABELS`` (no → desc → unit → qty → rate
    → total), which means columns that are ambiguous fall through
    to the more "specific" interpretation.
    """
    h = _clean(header_text)
    out: Dict[str, int] = {}
    for key, kws in _HEADER_LABELS.items():
        # Only count ONE match per label so a column with two qty-
        # related words doesn't double-count. We just want a binary
        # "is this column a candidate for label X" answer.
        for kw in kws:
            if kw.lower() in h:
                out[key] = 1
                break
    return out


def detect_columns(header_cells: List[Any]) -> Dict[str, int]:
    """Given a list of header cell values (one per column), return
    a mapping of logical column name to 0-indexed column offset.

    Strategy: score every header against every label, then greedily
    assign each column to its highest-scoring label, skipping labels
    already taken. The order of preference for ties is the
    ``_HEADER_LABELS`` dict order (no first, then desc, etc.).

    Returns a partial mapping — a column may not match any label and
    gets ignored.
    """
    candidates: List[Tuple[int, str, int]] = []  # (col_idx, label, score)
    for c_idx, val in enumerate(header_cells):
        scores = _score_column(val)
        for label, score in scores.items():
            # Each column-label pair gets score 1 if it matches at
            # all (we don't double-count within a column). The
            # scoring is "first match wins" per label-key.
            candidates.append((c_idx, label, score))

    # Greedy: sort by score desc, then by label priority (no, desc,
    # unit, qty, rate, total) so the higher-priority label claims
    # a column first if multiple labels match it.
    label_priority = {key: i for i, key in enumerate(_HEADER_LABELS.keys())}
    candidates.sort(key=lambda t: (-t[2], label_priority.get(t[1], 99), t[0]))

    taken_cols: set[int] = set()
    taken_labels: set[str] = set()
    result: Dict[str, int] = {}
    for c_idx, label, _ in candidates:
        if c_idx in taken_cols or label in taken_labels:
            continue
        result[label] = c_idx
        taken_cols.add(c_idx)
        taken_labels.add(label)
    return result


def parse_excel(file_path: str) -> Tuple[str, Dict[str, Any], Dict[str, Any]]:
    """Parse an input Excel file, returning markdown + data + headers.

    The output ``data_mapping`` uses the canonical English key names
    (``"Nr."``, ``"Item Description"``, ``"Unit"``, ``"Qty"``, ``"Rate"``,
    ``"Amount"``) so the Excel writer can stay simple. Per-row data
    is also kept under ``original_values`` with the same English
    keys plus lowercased aliases for compatibility with the legacy
    test_core.py.

    Raises:
        ValueError: if the file is not a valid Excel workbook
            (corrupt zip, password-protected, wrong format, empty).
        FileNotFoundError: if the path doesn't exist.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(
            f"Excel file not found: {file_path}"
        )
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as e:
        # InvalidFileException = not a valid xlsx (could be .xls, .csv,
        # a password-protected file, or random bytes renamed to .xlsx).
        # BadZipFile = the file is corrupt or truncated.
        log.exception("parse_excel: invalid Excel file %s", file_path)
        raise ValueError(
            f"'{os.path.basename(file_path)}' is not a valid Excel file. "
            f"It may be password-protected, corrupt, or in an older format. "
            f"Re-export it from Excel as .xlsx and try again. "
            f"(Technical: {type(e).__name__}: {e})"
        ) from e
    except OSError as e:
        # PermissionError, file locked by Excel, network share offline.
        log.exception("parse_excel: OS error reading %s", file_path)
        if e.errno == errno.EACCES or isinstance(e, PermissionError):
            raise ValueError(
                f"Cannot read '{os.path.basename(file_path)}' — "
                f"the file is locked. Close it in Excel and try again."
            ) from e
        raise ValueError(
            f"Cannot read '{os.path.basename(file_path)}': {e}"
        ) from e

    markdown_parts: List[str] = []
    data_mapping: Dict[str, Any] = {}
    headers_mapping: Dict[str, Any] = {}

    global_id_counter = 1

    for sheet in wb.worksheets:
        # Find header row by scanning up to 50 rows.
        header_row_idx: Optional[int] = None
        mapped_cols: Dict[str, int] = {}

        for r_idx in range(1, min(51, sheet.max_row + 1)):
            row_vals = [sheet.cell(row=r_idx, column=c_idx).value
                        for c_idx in range(1, sheet.max_column + 1)]
            if all(v is None for v in row_vals):
                continue

            temp_map = detect_columns(row_vals)
            # Need at least one identifier (Nr OR Description) to
            # treat this row as a header. We don't require 2+ matches
            # anymore — sheets with just Nr+Description are valid.
            if "no" in temp_map or "desc" in temp_map:
                header_row_idx = r_idx
                mapped_cols = temp_map
                break

        if not header_row_idx:
            # Skip sheet if no header row was identified.
            continue

        # Collect pre-header metadata (project name, sheet title, etc.)
        pre_header_rows: List[str] = []
        for r_idx in range(1, header_row_idx):
            row_vals = [
                str(sheet.cell(row=r_idx, column=c).value).strip()
                for c in range(1, sheet.max_column + 1)
                if sheet.cell(row=r_idx, column=c).value is not None
            ]
            if row_vals:
                pre_header_rows.append(", ".join(row_vals))
        metadata_str = " : ".join(pre_header_rows) if pre_header_rows else ""

        # Record headers mapping (capitalised for the writer).
        headers_mapping[sheet.title] = {
            "Nr.": mapped_cols.get("no"),
            "Item Description": mapped_cols.get("desc"),
            "Unit": mapped_cols.get("unit"),
            "Qty": mapped_cols.get("qty"),
            "Rate": mapped_cols.get("rate"),
            "Amount": mapped_cols.get("total"),
            # Lowercase aliases for test_core.py compat.
            "nr": mapped_cols.get("no"),
            "description": mapped_cols.get("desc"),
            "unit": mapped_cols.get("unit"),
            "qty": mapped_cols.get("qty"),
            "rate": mapped_cols.get("rate"),
            "amount": mapped_cols.get("total"),
        }

        # Build markdown table header.
        sheet_md: List[str] = [f"## Worksheet: {sheet.title}"]
        if metadata_str:
            sheet_md.append(f"**Metadata**: {metadata_str}\n")
        sheet_md.append("| Global ID | Nr. | Item Description | Unit | Qty | Rate | Amount |")
        sheet_md.append("|---|---|---|---|---|---|---|")

        # Parse data rows.
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_cells = [sheet.cell(row=r_idx, column=c).value
                         for c in range(1, sheet.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in row_cells):
                continue

            def _cell(key: str) -> Any:
                idx = mapped_cols.get(key)
                if idx is None or idx >= len(row_cells):
                    return None
                return row_cells[idx]

            desc_val = _cell("desc")
            if desc_val is None or str(desc_val).strip() == "":
                # Skip rows that have no description — these are
                # sub-headers or visual spacers in real BOQs.
                continue

            nr = _cell("no")
            desc = str(desc_val).strip()
            unit = _cell("unit")
            qty = _cell("qty")
            rate = _cell("rate")
            amount = _cell("total")

            # Format Nr. — Arabic BOQs often use 1, 2, 3, ... so
            # stringify the int. Excel sometimes returns a float for
            # an integer cell; cast back if it's whole.
            if nr is None:
                nr_str = ""
            elif isinstance(nr, float) and nr.is_integer():
                nr_str = str(int(nr))
            else:
                nr_str = str(nr).strip()

            # Numeric coercion — if the cell holds a numeric value
            # we want a number, not a stringified version with
            # formatting artefacts. Strings that aren't parseable
            # become 0 so downstream formulas don't crash.
            def _to_num(v: Any) -> float:
                if v is None:
                    return 0.0
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    s = str(v).strip().replace(",", "").replace(" ", "")
                    return float(s)
                except (ValueError, TypeError):
                    return 0.0

            unit_str = str(unit).strip() if unit is not None else ""
            qty_num = _to_num(qty)
            rate_num = _to_num(rate)
            amount_num = _to_num(amount)

            g_id = f"R{global_id_counter}"
            global_id_counter += 1

            data_mapping[g_id] = {
                "Nr.": nr_str,
                "Item Description": desc,
                "Unit": unit_str,
                "Qty": qty_num,
                "Rate": rate_num,
                "Amount": amount_num,
                "sheet_name": sheet.title,
                "original_values": {
                    "Nr.": nr_str,
                    "nr": nr_str,
                    "Item Description": desc,
                    "description": desc,
                    "Unit": unit_str,
                    "unit": unit_str,
                    "Qty": qty_num,
                    "qty": qty_num,
                    "Rate": rate_num,
                    "rate": rate_num,
                    "Amount": amount_num,
                    "amount": amount_num,
                },
            }
            sheet_md.append(
                f"| {g_id} | {nr_str} | {desc} | {unit_str} | "
                f"{qty_num} | {rate_num} | {amount_num} |"
            )

        markdown_parts.append("\n".join(sheet_md))

    combined_markdown = "\n\n".join(markdown_parts)
    return combined_markdown, data_mapping, headers_mapping


def sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet name to remove invalid characters and limit length to 31 chars."""
    sanitized = re.sub(r'[\[\]\*\\\/\?\:]', '', name)
    sanitized = sanitized.strip()
    if not sanitized:
        sanitized = "General"
    # Pkg - <Name> has to be at most 31 characters.
    # "Pkg - " is 6 characters. So name can be at most 25 characters.
    return sanitized[:25].strip()


def _style_worksheet(
    ws,
    headers: List[str],
    items: List[Dict[str, Any]],
    sheet_title: str,
    desc_idx: int,
) -> None:
    """Apply the professional layout to one Pkg / Master sheet.

    Headers are bold-white on dark-slate, body cells use Calibri 11
    with wrap_text on the description column. Column widths are
    capped at the constants in this module so a single 800-char
    Arabic description doesn't break the layout. The header row
    is frozen so it stays visible during scroll. Alternating row
    stripes (zebra fill) make the body readable when the row
    count is large.
    """
    thin = _thin_border()
    header_font = Font(name=OUTPUT_FONT_NAME, size=HEADER_FONT_SIZE, bold=True, color=HEADER_FONT_HEX)
    header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    body_font = Font(name=OUTPUT_FONT_NAME, size=OUTPUT_FONT_SIZE, color=BODY_FONT_HEX)
    zebra_fill = PatternFill(start_color=ZEBRA_FILL_HEX, end_color=ZEBRA_FILL_HEX, fill_type="solid")

    # Column widths. We do the header first, then sample the body
    # so the cap actually has a chance to engage (the body of the
    # longest description, not the first row, is what was making
    # the column blow up to 800+ in the old version).
    col_caps = {
        1: COL_WIDTH_NR, 2: COL_WIDTH_DESC, 3: COL_WIDTH_UNIT,
        4: COL_WIDTH_QTY, 5: COL_WIDTH_RATE, 6: COL_WIDTH_AMOUNT,
        7: COL_WIDTH_SHEET, 8: COL_WIDTH_PACKAGE,
    }
    n_cols = max(len(headers), max((len(_to_list(it)) for it in items), default=0))
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        cap = col_caps.get(c, COL_WIDTH_DESC)
        if c == desc_idx:
            # Sample the LONGEST body value in this column, not the
            # header (header is short). Cap is what matters.
            longest = ""
            for it in items:
                vals = _to_list(it)
                if len(vals) >= c:
                    v = vals[c - 1]
                    if v is not None and len(str(v)) > len(longest):
                        longest = str(v)
            width = _column_width(longest, cap, pad=2)
        else:
            width = cap
        ws.column_dimensions[col_letter].width = width

    # Header row.
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    # Body rows. Apply borders, font, alignment, number format,
    # and zebra stripes.
    num_rows = len(items)
    for r_idx, item in enumerate(items, start=2):
        is_zebra = (r_idx % 2 == 0)
        vals = _to_list(item)
        for c_idx in range(1, n_cols + 1):
            v = vals[c_idx - 1] if c_idx - 1 < len(vals) else None
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = body_font
            cell.border = thin
            if is_zebra:
                cell.fill = zebra_fill

            # Alignment + wrap_text + number format per column role.
            if c_idx == 1:  # Nr
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif c_idx == 2:  # Item Description
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif c_idx == 3:  # Unit
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif c_idx in (4, 5):  # Qty, Rate
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                fmt = FMT_QTY if c_idx == 4 else FMT_RATE
                cell.number_format = fmt
            elif c_idx == 6:  # Amount — always a formula
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.value = f"=IFERROR(D{r_idx}*E{r_idx}, 0)"
                cell.number_format = FMT_AMOUNT
            else:  # Sheet / Package
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Page setup: fit to one page wide, repeat header on every printed page.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Tab color: cycle through a small palette so adjacent packages
    # are visually distinguishable when the user is scanning the tab bar.
    _TAB_PALETTE = [
        "FF2563EB",  # blue
        "FF059669",  # emerald
        "FFD97706",  # amber
        "FFDC2626",  # red
        "FF7C3AED",  # violet
        "FF0891B2",  # cyan
        "FFDB2777",  # pink
    ]
    try:
        ws.sheet_properties.tabColor = _TAB_PALETTE[
            sum(ord(c) for c in sheet_title) % len(_TAB_PALETTE)
        ]
    except Exception:
        pass


def _to_list(item: Any) -> List[Any]:
    """Normalise a row into a flat list of values for cell access.

    Items may be plain dicts (the parsed row shape) or anything
    that ``list()`` will flatten. We do NOT touch formulas — those
    are written by ``_style_worksheet`` after the cells exist.
    """
    if isinstance(item, dict):
        # Prefer original_values for the canonical keys, fall back
        # to the top-level dict.
        src = item.get("original_values") if "original_values" in item else item
        if not isinstance(src, dict):
            src = item
        # The standard 6 columns (in the same order as write_excel writes them).
        return [
            src.get("Nr.", src.get("nr", "")),
            src.get("Item Description", src.get("description", "")),
            src.get("Unit", src.get("unit", "")),
            src.get("Qty", src.get("qty", 0)),
            src.get("Rate", src.get("rate", 0)),
        ]
    if isinstance(item, (list, tuple)):
        return list(item)
    return [item]


def _style_cover(ws, project_name: str, date: str) -> None:
    """Render the Cover sheet: title + project name + date + summary block."""
    title_font = Font(name=OUTPUT_FONT_NAME, size=22, bold=True, color="FF1F2937")
    label_font = Font(name=OUTPUT_FONT_NAME, size=11, bold=True, color="FF374151")
    value_font = Font(name=OUTPUT_FONT_NAME, size=11, color="FF1F2937")
    thin = _thin_border()

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70

    ws["A1"] = "Tawreed"
    ws["A1"].font = Font(name=OUTPUT_FONT_NAME, size=26, bold=True, color="FF1F2937")
    ws["B1"] = "BOQ Work-Package Extractor"
    ws["B1"].font = Font(name=OUTPUT_FONT_NAME, size=14, color="FF6B7280", italic=True)

    ws.row_dimensions[1].height = 36

    # Metadata block.
    ws["A3"] = "Project Name"
    ws["A3"].font = label_font
    ws["B3"] = project_name or "—"
    ws["B3"].font = value_font
    if not project_name:
        ws["B3"].fill = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")

    ws["A4"] = "Date"
    ws["A4"].font = label_font
    ws["B4"] = date or "—"
    ws["B4"].font = value_font
    if not date:
        ws["B4"].fill = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")

    ws["A5"] = "Application"
    ws["A5"].font = label_font
    ws["B5"] = "Tawreed BOQ Processor v0.0.1"
    ws["B5"].font = value_font

    # Borders for the metadata block.
    for r in (3, 4, 5):
        for c in (1, 2):
            ws.cell(r, c).border = thin

    ws.freeze_panes = "A2"


def write_excel(
    output_path: str, row_mapping: dict, item_categories: dict,
    project_name: str, date: str, layout_style: str = "root",
) -> None:
    """Generate the deliverable Excel workbook.

    Layout (per Anthropic xlsx skill + user spec):
      1. Cover sheet (title + project + date + summary)
      2. One sheet per work package (Pkg - <name>)
      3. Master sheet (every item, with sheet + package columns)

    All cells use Calibri 11, the description column has wrap_text,
    every column has a hard width cap (60 for description per user
    spec), the header row is frozen, and the Amount column uses a
    ``=D*E`` formula with currency formatting.
    """
    wb = openpyxl.Workbook()

    # ---- Cover sheet -----------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "Cover"
    _style_cover(ws_cover, project_name, date)

    # ---- Package sheets --------------------------------------------------
    grouped_items: Dict[str, list] = {}
    for g_id, row_data in row_mapping.items():
        cat = item_categories.get(g_id, "General") or "General"
        grouped_items.setdefault(cat, []).append(row_data)

    created_sheets = {"Cover"}
    sheet_index = 1
    headers = ["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"]

    for cat_name, items in grouped_items.items():
        sanitized = sanitize_sheet_name(cat_name)
        base_title = f"Pkg - {sanitized}"
        sheet_title = base_title[:31]
        suffix_num = 1
        while sheet_title in created_sheets:
            suffix = f" ({suffix_num})"
            max_base_len = 31 - len(suffix)
            sheet_title = base_title[:max_base_len] + suffix
            suffix_num += 1
        created_sheets.add(sheet_title)

        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.showGridLines = False
        # _style_worksheet writes the header at row 1 and body at row 2+.
        # It also overwrites the Amount cell with =D*E* formula, so we
        # don't need to pre-append anything here.
        _style_worksheet(ws, headers, items, sheet_title, desc_idx=2)

    # ---- Master sheet ----------------------------------------------------
    master_title = "Master"
    ws_master = wb.create_sheet(title=master_title)
    ws_master.sheet_view.showGridLines = False

    has_sheet = any(
        isinstance(it, dict) and "sheet_name" in it
        for items in grouped_items.values() for it in items
    )
    master_headers = list(headers)
    if has_sheet:
        master_headers.append("Sheet")
    master_headers.append("Package")

    # Build flat list of master rows (raw values, formulas applied in style pass).
    master_rows: List[List[Any]] = []
    for cat_name, items in grouped_items.items():
        for item in items:
            vals = _to_list(item)
            row = [vals[0], vals[1], vals[2], vals[3], vals[4], 0]
            if has_sheet:
                sheet_name = item.get("sheet_name", "") if isinstance(item, dict) else ""
                row.append(sheet_name)
            row.append(cat_name)
            master_rows.append(row)

    _style_worksheet(ws_master, master_headers, master_rows, master_title, desc_idx=2)

    # ---- Save ------------------------------------------------------------
    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
    except PermissionError as e:
        # The most common cause: the user has the file open in Excel,
        # which takes an exclusive write lock on Windows.
        log.exception("write_excel: permission denied writing %s", output_path)
        raise IOError(
            f"Cannot write '{os.path.basename(output_path)}' — "
            f"the file is open in Excel or another program has it locked. "
            f"Close it and try again."
        ) from e
    except OSError as e:
        # Disk full, path too long, network share offline, etc.
        log.exception("write_excel: OS error writing %s", output_path)
        raise IOError(
            f"Cannot write '{os.path.basename(output_path)}': {e}"
        ) from e


def parse_excel_boq(file_path: str) -> tuple[str, dict, dict]:
    """Wrapper function for compatibility with root test_core.py."""
    return parse_excel(file_path)


def generate_output_excel(project_name: str, date: str, row_mapping: dict, item_categories: dict, output_path: str) -> None:
    """Wrapper function for compatibility with root test_core.py."""
    return write_excel(output_path, row_mapping, item_categories, project_name, date)
