import os
import json
import sqlite3
import unittest
import shutil
import openpyxl
from openpyxl.styles import PatternFill

# Adjust import path
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from core.db import init_db, get_history, add_history, get_settings, save_settings, get_outputs_dir, DB_PATH, CONFIG_PATH, TAWREED_DIR
from core.excel import parse_excel, write_excel

class TestCoreTawreed(unittest.TestCase):
    
    def setUp(self):
        # Ensure directories exist
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        
        # Backup existing configurations and db if any
        self.db_backup = DB_PATH + ".bak" if os.path.exists(DB_PATH) else None
        if self.db_backup:
            shutil.copy2(DB_PATH, self.db_backup)
            try:
                os.remove(DB_PATH)
            except Exception:
                pass
            
        self.config_backup = CONFIG_PATH + ".bak" if os.path.exists(CONFIG_PATH) else None
        if self.config_backup:
            shutil.copy2(CONFIG_PATH, self.config_backup)
            try:
                os.remove(CONFIG_PATH)
            except Exception:
                pass
                
        # Initialize fresh DB
        init_db()

    def tearDown(self):
        # Clean up database connection if any is open
        # Restore backups
        if self.db_backup and os.path.exists(self.db_backup):
            try:
                if os.path.exists(DB_PATH):
                    os.remove(DB_PATH)
                shutil.copy2(self.db_backup, DB_PATH)
                os.remove(self.db_backup)
            except Exception:
                pass
            
        if self.config_backup and os.path.exists(self.config_backup):
            try:
                if os.path.exists(CONFIG_PATH):
                    os.remove(CONFIG_PATH)
                shutil.copy2(self.config_backup, CONFIG_PATH)
                os.remove(self.config_backup)
            except Exception:
                pass
            
        # Clean up test temp files
        for f in ["test_input.xlsx", "test_output.xlsx", "test_output_missing.xlsx"]:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except Exception:
                    pass

    def test_database_and_settings(self):
        print("\n--- Testing Database & Settings ---")
        # 1. Test settings
        default_s = get_settings()
        self.assertEqual(default_s["model_id"], "MiniMax-M3")
        
        test_settings = {
            "api_key": "sk-test12345",
            "model_id": "test-model-id",
            "base_url": "https://api.test-server.com/v1"
        }
        save_settings(test_settings)
        saved_s = get_settings()
        self.assertEqual(saved_s["api_key"], "sk-test12345")
        self.assertEqual(saved_s["model_id"], "test-model-id")
        self.assertEqual(saved_s["base_url"], "https://api.test-server.com/v1")
        
        # 2. Test history
        history_before = get_history()
        # Verify db runs in WAL mode
        conn = sqlite3.connect(DB_PATH)
        try:
            journal_mode = conn.execute("PRAGMA journal_mode").fetchone()[0]
            self.assertEqual(journal_mode.lower(), "wal")
        finally:
            conn.close()
            
        add_history("Test Project Alpha", 3, "outputs/alpha.xlsx")
        add_history("Test Project Beta", 4, "outputs/beta.xlsx")
        
        history_after = get_history()
        self.assertEqual(len(history_after), len(history_before) + 2)
        
        # Records sorted by ID DESC
        self.assertEqual(history_after[0]["project_name"], "Test Project Beta")
        self.assertEqual(history_after[0]["packages_count"], 4)
        self.assertEqual(history_after[0]["output_path"], "outputs/beta.xlsx")
        self.assertIn("timestamp", history_after[0])
        
        self.assertEqual(history_after[1]["project_name"], "Test Project Alpha")
        self.assertEqual(history_after[1]["packages_count"], 3)
        self.assertEqual(history_after[1]["output_path"], "outputs/alpha.xlsx")
        
        # Test outputs directory
        outputs_dir = get_outputs_dir()
        self.assertTrue(os.path.isabs(outputs_dir))
        self.assertTrue(os.path.exists(outputs_dir))
        print("Database & Settings tests PASSED.")

    def test_excel_parsing_and_writing(self):
        print("\n--- Testing Excel Parsing & Writing ---")
        # 1. Create a sample Excel sheet programmatically
        wb_in = openpyxl.Workbook()
        ws_in = wb_in.active
        ws_in.title = "Sheet1"
        
        # Pre-header rows
        ws_in['A1'] = "Sheet Title: My Sample Sheet, Page: 1"
        ws_in['A2'] = "Location: Sector B"
        
        # Header row with Arabic and Legacy garbled keywords
        ws_in.append([]) # Empty row 3
        # Row 4 is headers
        ws_in.append(["مسلسل", "بيان بند", "وحدة", "┘â┘à┘è╪⌐", "سعر", "اجمالي"])
        
        # Data rows
        ws_in.append(["1.1", "Excavation in sand", "m3", 100, 25, 2500])
        ws_in.append(["1.2", "Reinforced concrete foundations", "m3", 50, 200, 10000])
        ws_in.append(["1.3", "Plain concrete blinds", "m3", 20, 150, 3000])
        
        wb_in.save("test_input.xlsx")
        
        # 2. Parse Excel
        markdown_content, data_mapping, headers_mapping = parse_excel("test_input.xlsx")
        
        # Verify headers mapping
        self.assertIn("Sheet1", headers_mapping)
        h_map = headers_mapping["Sheet1"]
        self.assertEqual(h_map["Nr."], 0) # مسلسل
        self.assertEqual(h_map["Item Description"], 1) # بيان بند
        self.assertEqual(h_map["Unit"], 2) # وحدة
        self.assertEqual(h_map["Qty"], 3) # ┘â┘à┘è╪⌐
        self.assertEqual(h_map["Rate"], 4) # سعر
        self.assertEqual(h_map["Amount"], 5) # اجمالي
        
        # Verify metadata extraction (joined by " : ")
        self.assertIn("Sheet Title: My Sample Sheet, Page: 1 : Location: Sector B", markdown_content)
        
        # Verify markdown tables content
        self.assertIn("Excavation in sand", markdown_content)
        self.assertIn("Reinforced concrete foundations", markdown_content)
        self.assertIn("R1", markdown_content)
        self.assertIn("R2", markdown_content)
        self.assertIn("R3", markdown_content)
        
        # Verify data_mapping dictionary
        self.assertEqual(len(data_mapping), 3)
        self.assertEqual(data_mapping["R1"]["Item Description"], "Excavation in sand")
        self.assertEqual(data_mapping["R1"]["Qty"], 100)
        self.assertEqual(data_mapping["R1"]["Rate"], 25)
        self.assertEqual(data_mapping["R1"]["sheet_name"], "Sheet1")
        
        self.assertEqual(data_mapping["R2"]["Item Description"], "Reinforced concrete foundations")
        self.assertEqual(data_mapping["R2"]["Qty"], 50)
        
        # 3. Categorize items and write Excel
        item_categories = {
            "R1": "Excavation & Earthworks",
            "R2": "Concrete Works",
            "R3": "Concrete Works"
        }
        
        write_excel("test_output.xlsx", data_mapping, item_categories, "Sample Project Test", "2026-06-12", layout_style="core")
        
        # 4. Verify output file structure & content
        wb_out = openpyxl.load_workbook("test_output.xlsx")
        sheet_names = wb_out.sheetnames
        
        self.assertIn("Cover", sheet_names)
        self.assertIn("Pkg - Excavation & Earthworks", sheet_names) # Limit name length
        self.assertIn("Pkg - Concrete Works", sheet_names)
        self.assertIn("Master", sheet_names)
        
        # Check Cover details
        ws_cover = wb_out["Cover"]
        self.assertEqual(ws_cover["B3"].value, "Sample Project Test")
        self.assertEqual(ws_cover["B4"].value, "2026-06-12")
        
        # Check missing project name highlighting
        write_excel("test_output_missing.xlsx", data_mapping, item_categories, None, None, layout_style="core")
        wb_missing = openpyxl.load_workbook("test_output_missing.xlsx")
        ws_cover_missing = wb_missing["Cover"]
        self.assertIn(ws_cover_missing["B3"].fill.start_color.rgb, ["00FFFF00", "FFFF00"])
        self.assertIn(ws_cover_missing["B4"].fill.start_color.rgb, ["00FFFF00", "FFFF00"])
        
        # Check individual package sheet (Pkg - Concrete Works)
        ws_conc = wb_out["Pkg - Concrete Works"]
        # gridlines should be enabled
        self.assertTrue(ws_conc.views.sheetView[0].showGridLines)
        
        # It should contain table
        self.assertEqual(len(ws_conc.tables), 1)
        table_name = list(ws_conc.tables.keys())[0]
        self.assertTrue(table_name.startswith("Table_Pkg_"))
        
        # Verify headers
        headers = [cell.value for cell in ws_conc[1]]
        self.assertEqual(headers, ["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
        
        # Verify data rows & formula
        row2 = [cell.value for cell in ws_conc[2]]
        self.assertEqual(row2[0], "1.2")
        self.assertEqual(row2[1], "Reinforced concrete foundations")
        self.assertEqual(row2[2], "m3")
        self.assertEqual(row2[3], 50)
        self.assertEqual(row2[4], 200)
        self.assertEqual(row2[5], "=IFERROR(D2*E2, 0)")
        
        # Check Master sheet
        ws_master = wb_out["Master"]
        self.assertTrue(ws_master.views.sheetView[0].showGridLines)
        self.assertEqual(len(ws_master.tables), 1)
        table_name = list(ws_master.tables.keys())[0]
        self.assertEqual(table_name, "Table_Master")
        
        master_headers = [cell.value for cell in ws_master[1]]
        self.assertIn("Sheet", master_headers)
        self.assertEqual(master_headers, ["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount", "Sheet", "Package"])
        
        master_row3 = [cell.value for cell in ws_master[3]]
        self.assertEqual(master_row3[1], "Reinforced concrete foundations")
        self.assertEqual(master_row3[5], "=IFERROR(D3*E3, 0)")
        self.assertEqual(master_row3[6], "Sheet1")
        self.assertEqual(master_row3[7], "Concrete Works")
        
        print("Excel parsing & writing tests PASSED.")

if __name__ == "__main__":
    unittest.main()
