import os
import sys
import json
import sqlite3
import openpyxl
import openai
from datetime import datetime
from unittest.mock import MagicMock

# Ensure the workspace root is in python path
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

from core.db import (
    init_db, get_settings, save_settings, add_history, get_history,
    get_outputs_dir, CONFIG_PATH, DB_PATH
)
from core.excel import parse_excel, write_excel
from core.ai import analyze_boq_stream
import core.ai as ai_module

def test_db_operations():
    print("Running database & settings tests...")
    
    # 1. Backup existing config
    original_config = None
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                original_config = f.read()
        except Exception as e:
            print(f"Warning: Could not backup config: {e}")
            
    try:
        # Initialize DB (creates directories and db)
        init_db()
        assert os.path.exists(os.path.dirname(DB_PATH)), "DB directory was not created"
        assert os.path.exists(os.path.dirname(CONFIG_PATH)), "Config directory was not created"
        
        # Test outputs directory
        outputs_dir = get_outputs_dir()
        assert os.path.exists(outputs_dir), "Outputs directory was not created"
        assert os.path.isabs(outputs_dir), "Outputs directory path is not absolute"
        
        # Test settings saving and loading
        test_settings = {
            "api_key": "test_api_key_12345",
            "model_id": "test-model-id",
            "base_url": "https://test.api.url/v1"
        }
        save_settings(test_settings)
        
        loaded = get_settings()
        assert loaded["api_key"] == "test_api_key_12345", "API Key mismatch"
        assert loaded["model_id"] == "test-model-id", "Model ID mismatch"
        assert loaded["base_url"] == "https://test.api.url/v1", "Base URL mismatch"
        
        # Test default settings fallback
        if os.path.exists(CONFIG_PATH):
            os.remove(CONFIG_PATH)
        defaults = get_settings()
        assert defaults["model_id"] == "MiniMax-M3", "Default model ID mismatch"
        assert defaults["base_url"] == "https://api.minimax.io/v1", "Default base URL mismatch"
        assert defaults["api_key"] == "", "Default API key should be empty"
        
        # Test history insert and fetch
        test_proj = f"Temp_Test_Proj_{int(datetime.now().timestamp())}"
        add_history(test_proj, 12, "dummy/path.xlsx")
        
        history = get_history()
        assert len(history) > 0, "History is empty after insert"
        
        # Verify first item (sorted by id DESC, so it should be our newly inserted one)
        latest = history[0]
        assert latest["project_name"] == test_proj, "History project name mismatch"
        assert latest["packages_count"] == 12, "History packages count mismatch"
        assert latest["output_path"] == "dummy/path.xlsx", "History output path mismatch"
        
        # Cleanup test history item from DB to keep it clean
        conn = sqlite3.connect(DB_PATH)
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM history WHERE project_name = ?", (test_proj,))
            conn.commit()
        finally:
            conn.close()
            
        print("Database & settings tests: PASSED")
        
    finally:
        # Restore original config.json
        if original_config is not None:
            try:
                os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
                with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
                    f.write(original_config)
            except Exception as e:
                print(f"Warning: Could not restore config: {e}")
        elif os.path.exists(CONFIG_PATH):
            os.remove(CONFIG_PATH)

def test_excel_operations():
    print("Running Excel parser and writer tests...")
    
    temp_input = "temp_test_input.xlsx"
    temp_output = "temp_test_output.xlsx"
    
    try:
        # 1. Create a dummy input Excel sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOQ Sheet 1"
        
        # Pre-header rows / metadata
        ws["A1"] = "Client"
        ws["B1"] = "Ministry of Housing"
        ws["A2"] = "Project Name"
        ws["B2"] = "Al-Naseem Residential Complex"
        
        # Header row at row 4
        headers = ["رقم", "بيان", "وحدة", "الكمية", "سعر", "اجمالي"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=4, column=col_idx, value=h)
            
        # Data rows
        data = [
            ["1.1", "Excavation in ordinary soil", "m3", 1200.0, 45.0, 54000.0],
            ["1.2", "Reinforced concrete in footings", "m3", 450.0, 850.0, 382500.0],
            ["2.1", "Hollow block masonry 20cm", "m2", 1500.0, 95.0, 142500.0],
            ["3.1", "Internal paint works", "m2", 4000.0, 30.0, 120000.0]
        ]
        
        for r_offset, row_vals in enumerate(data):
            r_idx = 5 + r_offset
            for col_idx, val in enumerate(row_vals, 1):
                ws.cell(row=r_idx, column=col_idx, value=val)
                
        wb.save(temp_input)
        
        # 2. Parse the dummy Excel sheet
        full_md, row_mapping, headers_map = parse_excel(temp_input)
        
        # Assertions on parser output
        assert "BOQ Sheet 1" in full_md, "Sheet name missing from markdown"
        assert "Al-Naseem Residential Complex" in full_md, "Metadata missing from markdown"
        assert "Excavation in ordinary soil" in full_md, "Data row missing from markdown"
        
        assert len(row_mapping) == 4, f"Expected 4 rows mapped, got {len(row_mapping)}"
        assert "R1" in row_mapping, "Global ID R1 missing from row mapping"
        
        r1_val = row_mapping["R1"]["original_values"]
        assert r1_val["nr"] == "1.1", "R1 Nr mismatch"
        assert r1_val["description"] == "Excavation in ordinary soil", "R1 Description mismatch"
        assert r1_val["qty"] == 1200.0, "R1 Qty mismatch"
        assert r1_val["rate"] == 45.0, "R1 Rate mismatch"
        
        assert "BOQ Sheet 1" in headers_map, "Sheet name missing from headers map"
        detected_cols = headers_map["BOQ Sheet 1"]
        assert detected_cols["nr"] == 0, "Nr column index mismatch"
        assert detected_cols["description"] == 1, "Description column index mismatch"
        assert detected_cols["qty"] == 3, "Qty column index mismatch"
        
        # 3. Write categorized Excel sheet
        item_categories = {
            "R1": "Earth Works",
            "R2": "Concrete Works",
            "R3": "Masonry Works",
            "R4": "Finishes"
        }
        
        write_excel(
            output_path=temp_output,
            row_mapping=row_mapping,
            item_categories=item_categories,
            project_name="Al-Naseem Residential Complex",
            date="2026-06-12"
        )
        
        # 4. Verify output Excel sheet
        assert os.path.exists(temp_output), "Output Excel file was not written"
        
        wb_out = openpyxl.load_workbook(temp_output, data_only=False)
        sheet_names = wb_out.sheetnames
        
        # Check sheet order and names
        assert "Cover" in sheet_names, "Cover sheet missing"
        assert "Master" in sheet_names, "Master sheet missing"
        assert "Pkg - Earth Works" in sheet_names, "Earth Works package sheet missing"
        assert "Pkg - Concrete Works" in sheet_names, "Concrete Works package sheet missing"
        
        # Check Cover values and formatting
        ws_cov = wb_out["Cover"]
        assert ws_cov["B1"].value == "Tawreed BOQ Processor", "Cover App mismatch"
        assert ws_cov["B2"].value == "Al-Naseem Residential Complex", "Cover Project mismatch"
        assert ws_cov["B3"].value == "2026-06-12", "Cover Date mismatch"
        
        # Check Master values
        ws_mas = wb_out["Master"]
        assert ws_mas.sheet_view.showGridLines, "Grid lines should be visible"
        # Table checks
        assert len(ws_mas._tables) == 1, "Master sheet should have 1 Table wrapper"
        # Formula check
        amount_cell_formula = ws_mas["F2"].value # Row 2 in Master corresponds to the first data row
        assert amount_cell_formula == "=IFERROR(D2*E2, 0)", f"Expected formula, got: {amount_cell_formula}"
        
        # Check Package sheets
        ws_pkg = wb_out["Pkg - Earth Works"]
        assert len(ws_pkg._tables) == 1, "Package sheet should have 1 Table wrapper"
        # Confirm column width was auto-fitted
        width = ws_pkg.column_dimensions['B'].width
        assert width > 10, f"Column B width not adjusted: {width}"
        
        print("Excel parser and writer tests: PASSED")
        
    finally:
        # Cleanup files
        if os.path.exists(temp_input):
            os.remove(temp_input)
        if os.path.exists(temp_output):
            os.remove(temp_output)

# Mock classes for OpenAI streaming completions
class MockDelta:
    def __init__(self, content=None, reasoning_content=None):
        self.content = content
        self.reasoning_content = reasoning_content

class MockChoice:
    def __init__(self, delta):
        self.delta = delta

class MockChunk:
    def __init__(self, delta):
        self.choices = [MockChoice(delta)]

def test_ai_streaming_parser():
    print("Running AI streaming parser tests...")
    
    # 1. Define dummy stream chunks
    chunks = [
        # DeepSeek R1-style thoughts via reasoning_content delta
        MockChunk(MockDelta(reasoning_content="Initial thoughts on ")),
        MockChunk(MockDelta(reasoning_content="categorization.")),
        # Custom tags inside content
        MockChunk(MockDelta(content="<think>Additional internal ")),
        MockChunk(MockDelta(content="reasoning </think>")),
        # JSON payload
        MockChunk(MockDelta(content='{\n  "project_name": "Tawreed test",\n')),
        MockChunk(MockDelta(content='  "date": "2026-06-12",\n')),
        MockChunk(MockDelta(content='  "items": {\n')),
        # Split tokens
        MockChunk(MockDelta(content='    "R1": "Concrete ')),
        MockChunk(MockDelta(content='Works",\n')),
        MockChunk(MockDelta(content='    "R2": "Finishes"\n')),
        MockChunk(MockDelta(content='  }\n}'))
    ]
    
    # Mock OpenAI client
    mock_client = MagicMock()
    mock_client.chat.completions.create.return_value = chunks
    
    # Patch openai.OpenAI
    original_client = openai.OpenAI
    openai.OpenAI = lambda **kwargs: mock_client
    
    try:
        # Invoke stream function
        gen = analyze_boq_stream(
            api_key="mock", base_url="mock", model_id="mock",
            system_prompt="mock", user_prompt="mock"
        )
        
        yielded = []
        parsed_result = None
        
        try:
            while True:
                token, is_thought = next(gen)
                yielded.append((token, is_thought))
        except StopIteration as e:
            parsed_result = e.value
            
        # Verify yielded tokens and thought states
        assert len(yielded) > 0, "No tokens yielded"
        
        # Verify thoughts were correctly marked
        thoughts = [tok for tok, is_th in yielded if is_th]
        contents = [tok for tok, is_th in yielded if not is_th]
        
        assert any("Initial thoughts" in t for t in thoughts), "DeepSeek reasoning thoughts missing"
        assert any("Additional internal" in t for t in thoughts), "In-content tag thoughts missing"
        
        # Check final output
        assert parsed_result is not None, "Parsed result was not returned"
        assert parsed_result["project_name"] == "Tawreed test", "Parsed project name mismatch"
        assert parsed_result["items"]["R1"] == "Concrete Works", "Parsed items R1 mismatch"
        assert parsed_result["items"]["R2"] == "Finishes", "Parsed items R2 mismatch"
        
        print("AI streaming parser tests: PASSED")
        
    finally:
        openai.OpenAI = original_client

if __name__ == "__main__":
    print("=================== STARTING TESTS ===================")
    try:
        test_db_operations()
        test_excel_operations()
        test_ai_streaming_parser()
        print("=================== ALL TESTS PASSED SUCCESSFULLY! ===================")
        sys.exit(0)
    except Exception as e:
        print(f"=================== TEST FAILURE: {e} ===================")
        import traceback
        traceback.print_exc()
        sys.exit(1)
